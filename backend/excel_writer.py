"""
Writes extracted sales data to an Excel file using openpyxl.
"""

from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


COLUMNS = [
    "timestamp",
    "sender",
    "product",
    "quantity",
    "unit_price",
    "total_price",
    "currency",
    "notes",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def write_to_excel(sales: List[Dict[str, Any]], output_path: str) -> str:
    """
    Write a list of sale records to an Excel workbook.

    Creates a single worksheet named "Sales" with a styled header row and
    auto-fitted column widths. Overwrites any existing file at output_path.

    Args:
        sales:       List of sale dicts. Expected keys match the COLUMNS list.
        output_path: Absolute path where the .xlsx file will be saved.

    Returns:
        The output_path string, confirming where the file was written.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Header row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, sale in enumerate(sales, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=sale.get(col_name, ""))

    # Auto-fit column widths
    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    wb.save(output_path)
    return output_path
